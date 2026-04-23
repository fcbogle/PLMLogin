"""Workbook writer and formatting for analysis outputs."""

from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from config import AppConfig
from src.models import AnalysisOutputs


HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E2F3")
PRODUCTION_TECHNICIAN_FILL = PatternFill(fill_type="solid", fgColor="DDEBF7")
DEDICATED_LICENCE_CANDIDATE_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")
CATEGORY_FILLS = {
    "Regular": PatternFill(fill_type="solid", fgColor="C6EFCE"),
    "Occasional": PatternFill(fill_type="solid", fgColor="FFEB9C"),
    "Rare": PatternFill(fill_type="solid", fgColor="F4CCCC"),
}
REPORT_CHART_HEIGHT = 10
REPORT_CHART_WIDTH = 18
REPORT_SECTION_SPACER_ROWS = 42
EXECUTIVE_TAB_COLOR = "70AD47"


def write_analysis_workbook(outputs: AnalysisOutputs, config: AppConfig) -> None:
    """Write all required sheets to a formatted Excel workbook."""

    output_path = config.output_file
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)

    create_overall_plm_analysis_sheet(workbook.create_sheet("A_Overall_PLM_Analysis"), outputs)
    create_named_licence_analysis_sheet(workbook.create_sheet("B_Named_License_Analysis"), outputs)
    create_adu_licence_analysis_sheet(workbook.create_sheet("C_ADU_License_Analysis"), outputs)
    create_reallocation_opportunity_sheet(workbook.create_sheet("D_Reallocation_Opportunity"), outputs)
    write_table_sheet(workbook.create_sheet("Named_Licence_Analysis"), outputs.named_licence_analysis)
    write_table_sheet(
        workbook.create_sheet("Licence_Recommendations"),
        outputs.licence_recommendations,
        category_column="usage_category",
    )
    write_table_sheet(
        workbook.create_sheet("Dedicated_Licence_Candidates"),
        outputs.dedicated_licence_candidates,
        category_column="usage_category",
    )
    create_usage_analysis_sheet(workbook.create_sheet("Usage_Analysis"), outputs)
    create_adu_analysis_sheet(workbook.create_sheet("ADU_Analysis"), outputs)
    write_table_sheet(
        workbook.create_sheet("Production_Tech_Review"),
        outputs.production_technician_matches,
    )
    write_table_sheet(
        workbook.create_sheet("User_Summary"),
        outputs.user_summary,
        category_column="usage_category",
        highlight_production_technicians=True,
        highlight_dedicated_licence_candidates=True,
        include_highlight_legend=True,
    )
    write_table_sheet(
        workbook.create_sheet("Regular_Users"),
        outputs.regular_users,
        category_column="usage_category",
        highlight_production_technicians=True,
        highlight_dedicated_licence_candidates=True,
    )
    write_table_sheet(
        workbook.create_sheet("Occasional_Users"),
        outputs.occasional_users,
        category_column="usage_category",
        highlight_production_technicians=True,
        highlight_dedicated_licence_candidates=True,
    )
    write_table_sheet(
        workbook.create_sheet("Rare_Users"),
        outputs.rare_users,
        category_column="usage_category",
        highlight_production_technicians=True,
        highlight_dedicated_licence_candidates=True,
    )
    create_reporting_sheet(workbook.create_sheet("Reporting"), outputs)
    write_table_sheet(workbook.create_sheet("ADU_Denied_Users"), outputs.adu_user_summary)
    write_table_sheet(workbook.create_sheet("Named_Licence_No_Login"), outputs.named_licence_no_login)
    write_table_sheet(workbook.create_sheet("Monthly_Usage"), outputs.monthly_activity)
    write_table_sheet(workbook.create_sheet("Monthly_ADU_Denials"), outputs.adu_monthly_denials)
    write_table_sheet(workbook.create_sheet("Raw_Login_Data"), outputs.raw_data)
    write_table_sheet(workbook.create_sheet("Raw_ADU_Data"), outputs.adu_raw_data)
    write_table_sheet(workbook.create_sheet("Source_Lists"), outputs.production_technician_matches)
    create_rules_and_assumptions_sheet(workbook.create_sheet("Rules_And_Assumptions"), outputs, config)

    for sheet_name in [
        "A_Overall_PLM_Analysis",
        "B_Named_License_Analysis",
        "C_ADU_License_Analysis",
        "D_Reallocation_Opportunity",
    ]:
        workbook[sheet_name].sheet_properties.tabColor = EXECUTIVE_TAB_COLOR

    workbook.save(output_path)


def create_overall_plm_analysis_sheet(worksheet: Worksheet, outputs: AnalysisOutputs) -> None:
    """Create executive tab A for overall PLM usage."""

    worksheet["A1"] = "A: Overall PLM Analysis"
    worksheet["A1"].font = Font(bold=True, size=16)
    worksheet["A2"] = "Source data: PLM login audit over the reporting period."

    overall_metrics = outputs.executive_summary.iloc[:5].reset_index(drop=True)
    estate_metrics = build_overall_estate_metrics(outputs)
    named_vs_adu_status = build_named_vs_adu_status_source(outputs)
    named_vs_average = build_named_vs_average_source(outputs)
    write_positioned_table(worksheet, overall_metrics, 4, 1, "PLM Usage Metrics")
    write_positioned_table(worksheet, outputs.category_summary, 4, 5, "PLM Users by Usage Category")
    write_positioned_table(worksheet, outputs.monthly_active_users, 4, 9, "Monthly PLM Users")
    write_positioned_table(worksheet, estate_metrics, 36, 1, "Licence Population Metrics")
    write_positioned_table(worksheet, named_vs_average, 36, 5, "Named Licence vs Usage Comparison")
    write_positioned_table(worksheet, named_vs_adu_status, 36, 9, "Named vs ADU Users by Status")

    add_category_split_chart(worksheet, "A16", 5, 5, len(outputs.category_summary))
    write_chart_explanation(
        worksheet,
        start_row=28,
        start_col=1,
        summary="Shows the split of users classed as Regular, Occasional, and Rare.",
        source="Built from the PLM login audit using average active login days per month.",
        message="Most PLM users are not regular users, which supports limiting dedicated licences to the most consistent users.",
        width_columns=6,
    )

    add_monthly_active_users_chart(worksheet, "H16", 9, 5, len(outputs.monthly_active_users))
    avg_monthly_users = round(outputs.monthly_active_users["active_users"].mean(), 2) if not outputs.monthly_active_users.empty else 0
    write_chart_explanation(
        worksheet,
        start_row=28,
        start_col=8,
        summary="Shows how many users accessed PLM in each reporting month.",
        source="Built from the PLM login audit using distinct users active in each month.",
        message=f"Average monthly PLM usage is {avg_monthly_users}, materially below total named licence allocation.",
        width_columns=6,
    )

    add_simple_bar_chart(
        worksheet,
        "A48",
        start_col=5,
        header_row=37,
        data_rows=len(named_vs_average),
        title="Named Licence Users vs Average Monthly PLM Users",
        category_col_offset=0,
        data_col_offset=1,
        x_axis_title="Metric",
        y_axis_title="Users",
        show_category_labels=False,
    )
    write_chart_explanation(
        worksheet,
        start_row=60,
        start_col=1,
        summary="Compares total named licence users with average monthly PLM users and named users with no observed login.",
        source="Built from the named licence assignment file and monthly distinct-user counts from the PLM login audit.",
        message="The named licence population is much larger than observed monthly usage, and a material subset of named users show no login activity.",
        width_columns=6,
    )

    add_stacked_status_chart(
        worksheet,
        "H48",
        start_col=9,
        header_row=37,
        data_rows=len(named_vs_adu_status),
        title="Named vs ADU Users by Status",
        x_axis_title="Assigned Group",
        y_axis_title="Users",
    )
    write_chart_explanation(
        worksheet,
        start_row=60,
        start_col=8,
        summary="Compares usage-status distribution across named users and ADU-assigned users.",
        source="Built from the named licence assignment file, PLM login audit, and ADU assignment subset.",
        message="Named and ADU populations show very different usage patterns, helping distinguish allocation issues from genuine shared-access demand.",
        width_columns=6,
    )
    style_dashboard_sheet(worksheet)


def create_named_licence_analysis_sheet(worksheet: Worksheet, outputs: AnalysisOutputs) -> None:
    """Create executive tab B for named licence analysis."""

    worksheet["A1"] = "B: Named License Analysis"
    worksheet["A1"].font = Font(bold=True, size=16)
    worksheet["A2"] = "Source data: named licence assignments matched to observed PLM usage."

    write_positioned_table(worksheet, outputs.named_licence_summary, 4, 1, "Named Licence Metrics")
    write_positioned_table(worksheet, outputs.named_licence_review_summary, 4, 5, "Named Licence Review Status")
    comparison_source = build_named_licence_allocation_chart_source(outputs)
    write_positioned_table(worksheet, comparison_source, 4, 9, "Named Licence Allocation Comparison")

    add_simple_bar_chart(
        worksheet,
        "A18",
        start_col=5,
        header_row=5,
        data_rows=len(outputs.named_licence_review_summary),
        title="Named Licence Usage Split",
        category_col_offset=0,
        data_col_offset=1,
        x_axis_title="Review Status",
        y_axis_title="Users",
        show_category_labels=False,
    )
    write_chart_explanation(
        worksheet,
        start_row=30,
        start_col=1,
        summary="Shows how named licence holders are distributed across retain, review, transfer, and no observed login statuses.",
        source="Built by matching the named licence assignment file to the PLM login audit.",
        message="A substantial share of named licences are assigned to users with low or no observed PLM usage.",
        width_columns=6,
    )

    add_simple_bar_chart(
        worksheet,
        "H18",
        start_col=9,
        header_row=5,
        data_rows=len(comparison_source),
        title="Named Licences Allocated vs Average Monthly Users",
        category_col_offset=0,
        data_col_offset=1,
        x_axis_title="Metric",
        y_axis_title="Users",
        show_category_labels=False,
    )
    write_chart_explanation(
        worksheet,
        start_row=30,
        start_col=8,
        summary="Compares the number of named licences allocated with average monthly PLM users.",
        source="Built from the named licence assignment file and average monthly distinct-user counts from the PLM login audit.",
        message="Named licence allocation exceeds observed monthly usage, indicating possible reassignment opportunities.",
        width_columns=6,
    )
    style_dashboard_sheet(worksheet)


def create_adu_licence_analysis_sheet(worksheet: Worksheet, outputs: AnalysisOutputs) -> None:
    """Create executive tab C for ADU licence pressure."""

    worksheet["A1"] = "C: ADU License Analysis"
    worksheet["A1"].font = Font(bold=True, size=16)
    worksheet["A2"] = "Source data: unsuccessful ADU/shared licence requests."

    adu_metrics = build_adu_executive_summary(outputs)
    threshold_info = build_executive_threshold_explanation()
    top_adu_users = build_ranked_adu_users_for_chart(outputs.adu_user_summary.head(20))

    write_positioned_table(worksheet, adu_metrics, 4, 1, "ADU Pressure Metrics")
    write_positioned_table(worksheet, threshold_info, 4, 5, "Dedicated Licence Threshold")
    write_positioned_table(worksheet, outputs.adu_monthly_denials, 4, 9, "ADU Denials by Month")
    write_positioned_table(worksheet, top_adu_users, 30, 1, "Users Most Affected by ADU Limits")
    write_positioned_table(worksheet, build_adu_user_label_lookup(top_adu_users), 30, 9, "ADU User Label Key")

    add_simple_bar_chart(
        worksheet,
        "A60",
        start_col=9,
        header_row=5,
        data_rows=len(outputs.adu_monthly_denials),
        title="ADU Licence Denials by Month",
        category_col_offset=0,
        data_col_offset=2,
        x_axis_title="Month",
        y_axis_title="Denied Attempts",
        show_category_labels=False,
    )
    write_chart_explanation(
        worksheet,
        start_row=72,
        start_col=1,
        summary="Shows monthly ADU/shared licence denial events across the reporting period.",
        source="Derived from the ADU insufficient-licence audit.",
        message="ADU demand is concentrated in specific periods and demonstrates access pressure in the shared licence model.",
        width_columns=6,
    )

    add_simple_bar_chart(
        worksheet,
        "H60",
        start_col=1,
        header_row=31,
        data_rows=len(top_adu_users),
        title="Users Most Affected by ADU Licence Limits",
        category_col_offset=0,
        data_col_offset=4,
        x_axis_title="User Label",
        y_axis_title="Denied Days",
        show_category_labels=False,
    )
    write_chart_explanation(
        worksheet,
        start_row=72,
        start_col=8,
        summary="Shows the ADU users with the highest number of denied days.",
        source="Derived from the ADU insufficient-licence audit and ranked by distinct denied days.",
        message="A small number of users account for the strongest evidence of unmet dedicated licence demand.",
        width_columns=6,
    )
    style_dashboard_sheet(worksheet)


def create_reallocation_opportunity_sheet(worksheet: Worksheet, outputs: AnalysisOutputs) -> None:
    """Create executive tab D for reallocation opportunity."""

    worksheet["A1"] = "D: Reallocation Opportunity"
    worksheet["A1"].font = Font(bold=True, size=16)
    worksheet["A2"] = (
        "Source data: combined evidence from PLM usage, ADU denials, Production Technicians, and named licences."
    )

    write_positioned_table(worksheet, outputs.licence_balance, 4, 1, "Licence Recovery vs Demand")
    write_positioned_table(worksheet, outputs.unused_licence_evidence, 4, 5, "Recoverable Licence Evidence")
    write_positioned_table(worksheet, outputs.recommendation_summary, 30, 1, "Recommendation Breakdown")
    write_positioned_table(worksheet, outputs.dedicated_licence_candidates, 30, 5, "Users Recommended for Dedicated Licences")

    add_simple_bar_chart(
        worksheet,
        "A60",
        start_col=1,
        header_row=5,
        data_rows=len(outputs.licence_balance),
        title="Licence Recovery vs Demand",
        category_col_offset=0,
        data_col_offset=1,
        x_axis_title="Licence Position",
        y_axis_title="Users",
        show_category_labels=False,
    )
    write_chart_explanation(
        worksheet,
        start_row=72,
        start_col=1,
        summary="Compares potential recoverable licences with high-confidence dedicated licence demand.",
        source="Built from named licence analysis, Production Technician review, and ADU candidate analysis.",
        message="Recoverable licence capacity appears sufficient to meet identified high-confidence demand.",
        width_columns=6,
    )

    add_simple_bar_chart(
        worksheet,
        "H60",
        start_col=1,
        header_row=31,
        data_rows=len(outputs.recommendation_summary),
        title="Recommendation Breakdown",
        category_col_offset=0,
        data_col_offset=1,
        x_axis_title="Recommendation",
        y_axis_title="Users",
        show_category_labels=False,
    )
    write_chart_explanation(
        worksheet,
        start_row=72,
        start_col=8,
        summary="Shows the count of users in each recommendation category.",
        source="Built from combined PLM usage, ADU denial, Production Technician, and named licence evidence.",
        message="The recommendation profile supports targeted reassignment rather than broad licence changes.",
        width_columns=6,
    )
    style_dashboard_sheet(worksheet)


def build_executive_threshold_explanation() -> pd.DataFrame:
    """Build executive-facing threshold explanation rows."""

    return pd.DataFrame(
        [
            {
                "threshold": "Dedicated licence candidate",
                "details": (
                    "Regular or Occasional PLM user with at least 10 distinct ADU denied days. "
                    "This threshold reflects the natural break between 19 and 7 denied days in the ADU data."
                ),
            },
            {
                "threshold": "Review candidate",
                "details": (
                    "Repeated ADU denial evidence below 10 denied days remains visible as Review licence need."
                ),
            },
        ]
    )


def build_named_licence_allocation_chart_source(outputs: AnalysisOutputs) -> pd.DataFrame:
    """Build a small comparison table for named licences versus monthly users."""

    named_licences = int(outputs.named_licence_summary.loc[
        outputs.named_licence_summary["metric"] == "Named licences allocated", "value"
    ].iloc[0]) if not outputs.named_licence_summary.empty else 0
    average_monthly_users = float(outputs.named_licence_summary.loc[
        outputs.named_licence_summary["metric"] == "Average monthly PLM users", "value"
    ].iloc[0]) if not outputs.named_licence_summary.empty else 0
    return pd.DataFrame(
        [
            {"metric": "Named licences allocated", "value": named_licences},
            {"metric": "Average monthly PLM users", "value": average_monthly_users},
        ]
    )


def build_adu_executive_summary(outputs: AnalysisOutputs) -> pd.DataFrame:
    """Build a compact ADU-focused executive table."""

    recommendation_counts = outputs.recommendation_summary.set_index("recommendation")["user_count"].to_dict()
    return pd.DataFrame(
        [
            {"metric": "Users with ADU denial evidence", "value": len(outputs.adu_user_summary)},
            {
                "metric": "High-confidence dedicated licence candidates",
                "value": recommendation_counts.get("Consider dedicated licence allocation", 0),
            },
            {
                "metric": "ADU users for review",
                "value": recommendation_counts.get("Review licence need", 0),
            },
            {
                "metric": "ADU users to monitor",
                "value": recommendation_counts.get("Monitor", 0),
            },
        ]
    )


def build_overall_estate_metrics(outputs: AnalysisOutputs) -> pd.DataFrame:
    """Build overall estate metrics for the top-level PLM tab."""

    named = outputs.named_licence_analysis
    avg_monthly_users = round(outputs.monthly_active_users["active_users"].mean(), 2) if not outputs.monthly_active_users.empty else 0
    adu_assigned = named.loc[named["allocated_licence"] == "PTC Navigate Contributor - ADU"] if not named.empty else named
    return pd.DataFrame(
        [
            {"metric": "Named licence users", "value": len(named)},
            {
                "metric": "Named licence users with no observed login",
                "value": int((named["match_status"] != "Matched").sum()) if not named.empty else 0,
            },
            {"metric": "ADU-assigned users", "value": len(adu_assigned)},
            {
                "metric": "ADU-assigned users with denial evidence",
                "value": len(outputs.adu_user_summary.merge(
                    adu_assigned[["named_user"]].rename(columns={"named_user": "user_display_name"}),
                    on="user_display_name",
                    how="inner",
                )) if not adu_assigned.empty and not outputs.adu_user_summary.empty else 0,
            },
            {
                "metric": "High-confidence dedicated ADU candidates",
                "value": int((outputs.licence_recommendations["recommendation"] == "Consider dedicated licence allocation").sum()),
            },
            {"metric": "Average monthly PLM users", "value": avg_monthly_users},
        ]
    )


def build_named_vs_average_source(outputs: AnalysisOutputs) -> pd.DataFrame:
    """Build a compact source table for named users versus average monthly usage."""

    named = outputs.named_licence_analysis
    avg_monthly_users = round(outputs.monthly_active_users["active_users"].mean(), 2) if not outputs.monthly_active_users.empty else 0
    no_login = int((named["match_status"] != "Matched").sum()) if not named.empty else 0
    return pd.DataFrame(
        [
            {"metric": "Named licence users", "value": len(named)},
            {"metric": "Average monthly PLM users", "value": avg_monthly_users},
            {"metric": "Named users with no observed login", "value": no_login},
        ]
    )


def build_named_vs_adu_status_source(outputs: AnalysisOutputs) -> pd.DataFrame:
    """Build comparison counts for named users versus ADU-assigned users by status."""

    statuses = ["Regular", "Occasional", "Rare", "No observed login"]
    named = outputs.named_licence_analysis.copy()
    adu_assigned = named.loc[named["allocated_licence"] == "PTC Navigate Contributor - ADU"].copy()
    named_status = named["usage_category"].fillna("No observed login")
    adu_status = adu_assigned["usage_category"].fillna("No observed login")
    rows = []
    for group_name, series in [("Named users", named_status), ("ADU-assigned users", adu_status)]:
        row = {"group": group_name}
        counts = series.value_counts().to_dict()
        for status in statuses:
            row[status] = int(counts.get(status, 0))
        rows.append(row)
    return pd.DataFrame(rows)


def create_usage_analysis_sheet(worksheet: Worksheet, outputs: AnalysisOutputs) -> None:
    """Create the usage analysis sheet with supporting trend data."""

    worksheet["A1"] = "PLM Usage Analysis"
    worksheet["A1"].font = Font(bold=True, size=16)
    write_positioned_table(worksheet, outputs.category_summary, 3, 1, "Usage Category Summary")
    write_positioned_table(worksheet, outputs.monthly_active_users, 3, 5, "Monthly PLM Users")
    write_positioned_table(worksheet, outputs.most_active_users, 3, 9, "Most Active Users")
    add_category_split_chart(worksheet, "A10", 1, 4, len(outputs.category_summary))
    add_monthly_active_users_chart(worksheet, "E20", 5, 4, len(outputs.monthly_active_users))
    format_sheet(worksheet)
    worksheet.sheet_view.showGridLines = False


def create_adu_analysis_sheet(worksheet: Worksheet, outputs: AnalysisOutputs) -> None:
    """Create the ADU analysis sheet with denial trends and ranked users."""

    worksheet["A1"] = "ADU Licence Denial Analysis"
    worksheet["A1"].font = Font(bold=True, size=16)
    top_adu_users = build_ranked_adu_users_for_chart(outputs.adu_user_summary.head(20))
    write_positioned_table(worksheet, outputs.adu_monthly_denials, 3, 1, "ADU Denials by Month")
    write_positioned_table(worksheet, top_adu_users, 3, 6, "Users Most Affected by ADU Limits")
    add_simple_bar_chart(
        worksheet,
        "A20",
        start_col=1,
        header_row=4,
        data_rows=len(outputs.adu_monthly_denials),
        title="ADU Licence Denials by Month",
        category_col_offset=0,
        data_col_offset=2,
        x_axis_title="Month",
        y_axis_title="Denied Attempts",
        show_category_labels=False,
    )
    add_simple_bar_chart(
        worksheet,
        "F28",
        start_col=6,
        header_row=4,
        data_rows=len(top_adu_users),
        title="Users Most Affected by ADU Licence Limits",
        category_col_offset=0,
        data_col_offset=4,
        x_axis_title="User Label",
        y_axis_title="Denied Days",
        show_category_labels=False,
    )
    write_positioned_table(
        worksheet,
        build_adu_user_label_lookup(top_adu_users),
        start_row=28,
        start_col=12,
        title="ADU User Label Key",
    )
    format_sheet(worksheet)
    worksheet.sheet_view.showGridLines = False


def build_ranked_adu_users_for_chart(adu_users: pd.DataFrame) -> pd.DataFrame:
    """Add simple U1, U2 labels for ranked ADU-user chart bins."""

    ranked = adu_users.copy().reset_index(drop=True)
    if ranked.empty:
        return ranked
    ranked.insert(0, "user_label", [f"U{index}" for index in range(1, len(ranked) + 1)])
    return ranked


def build_adu_user_label_lookup(ranked_adu_users: pd.DataFrame) -> pd.DataFrame:
    """Build a compact lookup table for U1, U2 ADU-user chart labels."""

    if ranked_adu_users.empty:
        return pd.DataFrame(columns=["user_label", "user_display_name"])
    return ranked_adu_users[["user_label", "user_display_name"]].copy()


def create_reporting_sheet(worksheet: Worksheet, outputs: AnalysisOutputs) -> None:
    """Create a reporting sheet with vertically stacked source tables and charts."""

    worksheet["A1"] = "PLM Usage and Licence Reporting"
    worksheet["A1"].font = Font(bold=True, size=16)

    current_row = 3
    current_row = write_reporting_section(worksheet, "Executive Metrics", outputs.executive_summary, current_row)
    current_row = write_reporting_section(
        worksheet,
        "Recoverable Licence Evidence",
        outputs.unused_licence_evidence,
        current_row,
    )
    current_row = write_reporting_section(
        worksheet,
        "Users Recommended for Dedicated Licences",
        outputs.dedicated_licence_candidates,
        current_row,
    )

    category_header_row = current_row + 1
    current_row = write_reporting_section(
        worksheet,
        "Usage Category Split Source Data",
        outputs.category_summary,
        current_row,
    )
    add_category_split_chart(
        worksheet=worksheet,
        anchor_cell=f"A{current_row}",
        start_col=1,
        header_row=category_header_row,
        data_rows=len(outputs.category_summary),
    )
    current_row += REPORT_SECTION_SPACER_ROWS

    recommendation_header_row = current_row + 1
    current_row = write_reporting_section(
        worksheet,
        "Recommendation Breakdown Source Data",
        outputs.recommendation_summary,
        current_row,
    )
    add_simple_bar_chart(
        worksheet=worksheet,
        anchor_cell=f"A{current_row}",
        start_col=1,
        header_row=recommendation_header_row,
        data_rows=len(outputs.recommendation_summary),
        title="Recommendation Breakdown",
        category_col_offset=0,
        data_col_offset=1,
        x_axis_title="Recommendation",
        y_axis_title="Users",
    )
    current_row += REPORT_SECTION_SPACER_ROWS

    monthly_header_row = current_row + 1
    current_row = write_reporting_section(
        worksheet,
        "Monthly PLM Users Source Data",
        outputs.monthly_active_users,
        current_row,
    )
    add_monthly_active_users_chart(
        worksheet=worksheet,
        anchor_cell=f"A{current_row}",
        start_col=1,
        header_row=monthly_header_row,
        data_rows=len(outputs.monthly_active_users),
    )
    current_row += REPORT_SECTION_SPACER_ROWS

    adu_monthly_header_row = current_row + 1
    current_row = write_reporting_section(
        worksheet,
        "ADU Denials by Month Source Data",
        outputs.adu_monthly_denials,
        current_row,
    )
    add_simple_bar_chart(
        worksheet=worksheet,
        anchor_cell=f"A{current_row}",
        start_col=1,
        header_row=adu_monthly_header_row,
        data_rows=len(outputs.adu_monthly_denials),
        title="ADU Licence Denials by Month",
        category_col_offset=0,
        data_col_offset=2,
        x_axis_title="Month",
        y_axis_title="Denied Attempts",
        show_category_labels=False,
    )
    current_row += REPORT_SECTION_SPACER_ROWS

    adu_users_header_row = current_row + 1
    top_adu_users = build_ranked_adu_users_for_chart(outputs.adu_user_summary.head(20))
    current_row = write_reporting_section(
        worksheet,
        "Users Most Affected by ADU Licence Limits",
        top_adu_users,
        current_row,
    )
    add_simple_bar_chart(
        worksheet=worksheet,
        anchor_cell=f"A{current_row}",
        start_col=1,
        header_row=adu_users_header_row,
        data_rows=len(top_adu_users),
        title="Users Most Affected by ADU Licence Limits",
        category_col_offset=0,
        data_col_offset=4,
        x_axis_title="User Label",
        y_axis_title="Denied Days",
        show_category_labels=False,
    )
    current_row += REPORT_SECTION_SPACER_ROWS

    production_header_row = current_row + 1
    current_row = write_reporting_section(
        worksheet,
        "Production Technician Match Review",
        outputs.production_technician_matches,
        current_row,
    )
    current_row += 4

    current_row = write_reporting_section(
        worksheet,
        "Licence Recovery vs Demand",
        outputs.licence_balance,
        current_row,
    )

    style_reporting_sheet(worksheet)


def create_overview_sheet(worksheet: Worksheet, outputs: AnalysisOutputs) -> None:
    """Create the overview sheet with metrics and category summary."""

    worksheet.append(["Metric", "Value"])
    for row in outputs.overview_metrics.itertuples(index=False):
        worksheet.append(list(row))

    worksheet["D1"] = "Category"
    worksheet["E1"] = "User Count"
    worksheet["F1"] = "Percentage of Total Users"
    for row_index, row in enumerate(outputs.category_summary.itertuples(index=False), start=2):
        worksheet.cell(row=row_index, column=4, value=row.category)
        worksheet.cell(row=row_index, column=5, value=row.user_count)
        worksheet.cell(row=row_index, column=6, value=row.percentage_of_total_users)

    format_sheet(worksheet)
    worksheet.freeze_panes = "A2"


def create_rules_and_assumptions_sheet(
    worksheet: Worksheet,
    outputs: AnalysisOutputs,
    config: AppConfig,
) -> None:
    """Create the rules and assumptions sheet."""

    assumptions = pd.DataFrame(
        [
            {
                "area": "Usage categorisation",
                "rule_or_assumption": (
                    "Regular, Occasional, and Rare are based on average distinct active login days per month."
                ),
            },
            {
                "area": "Regular threshold",
                "rule_or_assumption": (
                    f"Regular: average_active_days_per_month >= "
                    f"{config.thresholds.regular_min_average_active_days}"
                ),
            },
            {
                "area": "Occasional threshold",
                "rule_or_assumption": (
                    f"Occasional: average_active_days_per_month >= "
                    f"{config.thresholds.occasional_min_average_active_days} and below Regular threshold"
                ),
            },
            {
                "area": "Rare threshold",
                "rule_or_assumption": (
                    f"Rare: average_active_days_per_month < "
                    f"{config.thresholds.occasional_min_average_active_days}"
                ),
            },
            {
                "area": "ADU repeated-denial rule",
                "rule_or_assumption": (
                    f"Repeated ADU denial means at least {config.adu_repeated_denial_days_threshold} denied days "
                    f"or at least {config.adu_repeated_denial_attempts_threshold} denied attempts."
                ),
            },
            {
                "area": "Dedicated licence candidate rule",
                "rule_or_assumption": (
                    "Dedicated licence candidates must be Regular or Occasional PLM users with at least "
                    f"{config.dedicated_licence_denied_days_threshold} distinct ADU denied days."
                ),
            },
            {
                "area": "Production Technician matching",
                "rule_or_assumption": (
                    "Production Technicians are matched to the login audit by normalised full name."
                ),
            },
            {
                "area": "No observed authentication",
                "rule_or_assumption": (
                    "Production Technicians not matched to the PLM login audit after name matching are treated "
                    "as having no observed PLM authentication during the reporting period."
                ),
            },
        ]
    )
    worksheet["A1"] = "Rules and Assumptions"
    worksheet["A1"].font = Font(bold=True, size=16)
    write_positioned_table(worksheet, assumptions, 3, 1, "Analysis Rules")
    write_positioned_table(worksheet, outputs.category_rules, 14, 1, "Category Rules")
    format_sheet(worksheet)


def write_table_sheet(
    worksheet: Worksheet,
    dataframe,
    category_column: str | None = None,
    highlight_production_technicians: bool = False,
    highlight_dedicated_licence_candidates: bool = False,
    include_highlight_legend: bool = False,
) -> None:
    """Write a dataframe to a worksheet and apply basic formatting."""

    start_row = 1
    if include_highlight_legend:
        start_row = write_highlight_legend(worksheet)

    worksheet.append(list(dataframe.columns))
    for row in dataframe.itertuples(index=False, name=None):
        worksheet.append([excel_safe_value(value) for value in row])

    format_sheet(worksheet, header_row=start_row)
    if category_column and category_column in dataframe.columns:
        apply_category_highlighting(
            worksheet,
            dataframe.columns.get_loc(category_column) + 1,
            start_row + 1,
        )
    if highlight_production_technicians and "is_production_technician" in dataframe.columns:
        apply_production_technician_highlighting(
            worksheet,
            dataframe.columns.get_loc("is_production_technician") + 1,
            start_row + 1,
        )
    if highlight_dedicated_licence_candidates and "is_dedicated_licence_candidate" in dataframe.columns:
        apply_row_highlighting(
            worksheet,
            dataframe.columns.get_loc("is_dedicated_licence_candidate") + 1,
            DEDICATED_LICENCE_CANDIDATE_FILL,
            start_row + 1,
        )


def write_highlight_legend(worksheet: Worksheet) -> int:
    """Write a short legend explaining User Summary row highlighting."""

    worksheet["A1"] = "Highlight Legend"
    worksheet["A1"].font = Font(bold=True, size=12)
    worksheet["A2"] = "Light orange"
    worksheet["B2"] = "Recommended for dedicated licence allocation based on observed PLM usage and repeated ADU denials."
    worksheet["A2"].fill = DEDICATED_LICENCE_CANDIDATE_FILL
    worksheet["A3"] = "Light blue"
    worksheet["B3"] = "Matched Production Technician; highlighted for licence review context only."
    worksheet["A3"].fill = PRODUCTION_TECHNICIAN_FILL
    worksheet["A4"] = "Usage Category cell colours"
    worksheet["B4"] = "Regular, Occasional, and Rare values are conditionally coloured in the usage_category column."
    return 5


def write_positioned_table(
    worksheet: Worksheet,
    dataframe,
    start_row: int,
    start_col: int,
    title: str | None = None,
) -> int:
    """Write a dataframe at a specific location and return the next row."""

    row_index = start_row
    if title:
        worksheet.cell(row=row_index, column=start_col, value=title)
        worksheet.cell(row=row_index, column=start_col).font = Font(bold=True, size=12)
        row_index += 1

    for column_offset, header in enumerate(dataframe.columns):
        cell = worksheet.cell(row=row_index, column=start_col + column_offset, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    for data_row_offset, row in enumerate(dataframe.itertuples(index=False, name=None), start=1):
        for column_offset, value in enumerate(row):
            worksheet.cell(
                row=row_index + data_row_offset,
                column=start_col + column_offset,
                value=excel_safe_value(value),
            )
    return row_index + len(dataframe) + 2


def write_chart_explanation(
    worksheet: Worksheet,
    start_row: int,
    start_col: int,
    summary: str,
    source: str,
    message: str,
    width_columns: int = 6,
) -> int:
    """Write a standard chart explanation block."""

    entries = [
        ("Summary", summary),
        ("Source", source),
        ("Message", message),
    ]
    current_row = start_row
    for label, text in entries:
        worksheet.cell(row=current_row, column=start_col, value=f"{label}:")
        worksheet.cell(row=current_row, column=start_col).font = Font(bold=True)
        worksheet.cell(row=current_row, column=start_col + 1, value=text)
        current_row += 1

    end_col = start_col + width_columns - 1
    for column_index in range(start_col + 1, end_col + 1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = 22
    return current_row


def write_reporting_section(
    worksheet: Worksheet,
    title: str,
    dataframe,
    start_row: int,
) -> int:
    """Write a reporting section vertically and return the next available row."""

    worksheet.cell(row=start_row, column=1, value=title)
    worksheet.cell(row=start_row, column=1).font = Font(bold=True, size=12)

    header_row = start_row + 1
    for column_index, header in enumerate(dataframe.columns, start=1):
        cell = worksheet.cell(row=header_row, column=column_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    for row_offset, row in enumerate(dataframe.itertuples(index=False, name=None), start=1):
        for column_index, value in enumerate(row, start=1):
            worksheet.cell(row=header_row + row_offset, column=column_index, value=excel_safe_value(value))

    return header_row + len(dataframe) + 3


def excel_safe_value(value):
    """Convert pandas missing values to blank Excel cells."""

    if pd.isna(value):
        return None
    return value


def format_sheet(worksheet: Worksheet, header_row: int = 1) -> None:
    """Apply standard workbook formatting."""

    bold_font = Font(bold=True)
    for cell in worksheet[header_row]:
        cell.font = bold_font
        cell.fill = HEADER_FILL

    worksheet.freeze_panes = f"A{header_row + 1}"
    last_column = get_column_letter(worksheet.max_column)
    worksheet.auto_filter.ref = f"A{header_row}:{last_column}{worksheet.max_row}"

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len("" if cell.value is None else str(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 40)


def style_dashboard_sheet(worksheet: Worksheet) -> None:
    """Apply formatting for executive dashboard-style tabs."""

    worksheet.freeze_panes = "A3"
    worksheet.sheet_view.showGridLines = False
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len("" if cell.value is None else str(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 32)


def style_reporting_sheet(worksheet: Worksheet) -> None:
    """Apply formatting specific to the reporting sheet."""

    worksheet.freeze_panes = "A4"
    worksheet.sheet_view.showGridLines = False

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len("" if cell.value is None else str(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 32)


def apply_category_highlighting(
    worksheet: Worksheet,
    category_column_index: int,
    first_data_row: int = 2,
) -> None:
    """Apply simple fill colours to usage categories."""

    if worksheet.max_row < first_data_row:
        return

    column_letter = get_column_letter(category_column_index)
    data_range = f"{column_letter}{first_data_row}:{column_letter}{worksheet.max_row}"
    for category, fill in CATEGORY_FILLS.items():
        worksheet.conditional_formatting.add(
            data_range,
            CellIsRule(operator="equal", formula=[f'"{category}"'], fill=fill),
        )


def apply_production_technician_highlighting(
    worksheet: Worksheet,
    production_technician_column_index: int,
    first_data_row: int = 2,
) -> None:
    """Highlight rows matched to the Production Technician list."""

    apply_row_highlighting(
        worksheet,
        production_technician_column_index,
        PRODUCTION_TECHNICIAN_FILL,
        first_data_row,
    )


def apply_row_highlighting(
    worksheet: Worksheet,
    flag_column_index: int,
    fill: PatternFill,
    first_data_row: int = 2,
) -> None:
    """Highlight full rows where a boolean flag column is true."""

    if worksheet.max_row < first_data_row:
        return

    for row_index in range(first_data_row, worksheet.max_row + 1):
        if worksheet.cell(row=row_index, column=flag_column_index).value is True:
            for column_index in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row_index, column=column_index).fill = fill


def add_category_split_chart(
    worksheet: Worksheet,
    anchor_cell: str,
    start_col: int,
    header_row: int,
    data_rows: int,
) -> None:
    """Add a doughnut chart for user category split."""

    if data_rows < 1:
        return

    chart = DoughnutChart()
    chart.title = "Usage Category Split"
    chart.style = 10
    chart.height = REPORT_CHART_HEIGHT
    chart.width = REPORT_CHART_WIDTH

    labels = Reference(worksheet, min_col=start_col, min_row=header_row + 1, max_row=header_row + data_rows)
    data = Reference(worksheet, min_col=start_col + 1, min_row=header_row, max_row=header_row + data_rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showVal = True
    chart.dataLabels.showLeaderLines = True

    worksheet.add_chart(chart, anchor_cell)


def add_monthly_active_users_chart(
    worksheet: Worksheet,
    anchor_cell: str,
    start_col: int,
    header_row: int,
    data_rows: int,
) -> None:
    """Add a column chart for monthly active users."""

    if data_rows < 1:
        return

    chart = BarChart()
    chart.type = "col"
    chart.title = "Monthly PLM Users by Month"
    chart.y_axis.title = "Users"
    chart.x_axis.title = "Month Label"
    chart.y_axis.delete = False
    chart.x_axis.delete = False
    chart.style = 10
    chart.height = REPORT_CHART_HEIGHT
    chart.width = REPORT_CHART_WIDTH
    chart.legend = None
    chart.gapWidth = 35
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.showSerName = False
    chart.dataLabels.showCatName = False
    chart.dataLabels.showLegendKey = False

    data = Reference(worksheet, min_col=start_col + 2, min_row=header_row, max_row=header_row + data_rows)
    categories = Reference(worksheet, min_col=start_col, min_row=header_row + 1, max_row=header_row + data_rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    worksheet.add_chart(chart, anchor_cell)


def add_simple_bar_chart(
    worksheet: Worksheet,
    anchor_cell: str,
    start_col: int,
    header_row: int,
    data_rows: int,
    title: str,
    category_col_offset: int,
    data_col_offset: int,
    x_axis_title: str,
    y_axis_title: str,
    show_category_labels: bool = True,
) -> None:
    """Add a simple column chart from a positioned source table."""

    if data_rows < 1:
        return

    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.x_axis.title = x_axis_title
    chart.y_axis.title = y_axis_title
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.delete = False
    chart.x_axis.delete = False
    chart.style = 10
    chart.height = REPORT_CHART_HEIGHT
    chart.width = REPORT_CHART_WIDTH
    chart.legend = None
    chart.gapWidth = 35
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.showCatName = show_category_labels
    chart.dataLabels.showSerName = False
    chart.dataLabels.showLegendKey = False

    data = Reference(
        worksheet,
        min_col=start_col + data_col_offset,
        min_row=header_row,
        max_row=header_row + data_rows,
    )
    categories = Reference(
        worksheet,
        min_col=start_col + category_col_offset,
        min_row=header_row + 1,
        max_row=header_row + data_rows,
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    worksheet.add_chart(chart, anchor_cell)


def add_stacked_status_chart(
    worksheet: Worksheet,
    anchor_cell: str,
    start_col: int,
    header_row: int,
    data_rows: int,
    title: str,
    x_axis_title: str,
    y_axis_title: str,
) -> None:
    """Add a stacked column chart for status comparisons."""

    if data_rows < 1:
        return

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = title
    chart.x_axis.title = x_axis_title
    chart.y_axis.title = y_axis_title
    chart.style = 10
    chart.height = REPORT_CHART_HEIGHT
    chart.width = REPORT_CHART_WIDTH
    chart.gapWidth = 35

    data = Reference(
        worksheet,
        min_col=start_col + 1,
        min_row=header_row,
        max_col=start_col + 4,
        max_row=header_row + data_rows,
    )
    categories = Reference(
        worksheet,
        min_col=start_col,
        min_row=header_row + 1,
        max_row=header_row + data_rows,
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    worksheet.add_chart(chart, anchor_cell)


def add_ranked_users_chart(
    worksheet: Worksheet,
    anchor_cell: str,
    start_col: int,
    header_row: int,
    data_rows: int,
    title: str,
) -> None:
    """Add a horizontal bar chart for ranked user activity."""

    if data_rows < 1:
        return

    chart = BarChart()
    chart.type = "bar"
    chart.style = 12
    chart.title = title
    chart.y_axis.title = "User Name"
    chart.x_axis.title = "Average Active Days Per Month"
    chart.height = REPORT_CHART_HEIGHT
    chart.width = REPORT_CHART_WIDTH
    chart.legend = None
    chart.gapWidth = 40

    data = Reference(worksheet, min_col=start_col + 2, min_row=header_row, max_row=header_row + data_rows)
    categories = Reference(worksheet, min_col=start_col, min_row=header_row + 1, max_row=header_row + data_rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    worksheet.add_chart(chart, anchor_cell)
